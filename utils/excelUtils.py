import openpyxl

'''
title:文件名
path:文件保存目录
value:二维数组
'''
# 行
i = 0
# 列
j = 0


def write07Excel(title, path, value):
    global i, j
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = title

    for i in range(i, len(value[0])):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('2007测试表')

    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()
